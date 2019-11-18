# head
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image

import tkinter as tk
from tkinter import filedialog

from os import getcwd
from time import sleep, clock
from datetime import datetime
import warnings

def db_filler():

	# открываем заявку
	root = tk.Tk()
	zajavka = filedialog.askopenfilename()
	sleep(1)
	root.destroy()
	if len(zajavka) == 0:
		return print("Выбор файла отменён")
	wb = load_workbook(filename=zajavka, data_only=True)
	ws = wb.active

	# считаем позиции в заявке
	r = 13
	posCount = 0
	while type(ws['A{}'.format(r)].value) == int:
		posCount += 1
		r += 1

	# пытаемся открыть базу и дописать заявку
	try:
		wbr = load_workbook(filename='bdz.xlsx', data_only=True)
		wsr = wbr.active
		last_id = wsr['A'][-1].value

		date = ws['A7'].value
		year = '20' + date[-3:-1]
		r = 13
		id = last_id +1
		theme = 		ws['A9'].value
		makerD = 	ws['A{}'.format(18+posCount)].value
		makerFio = 	ws['F{}'.format(18+posCount)].value
		dopInfo = 	ws['A{}'.format(14+posCount)].value
		tupleAllZ = wsr['B'][1:]
		listZ = []

		for i in tupleAllZ:
			if i.value[-4:] == year:
				listZ.append(int(i.value[0:-5]))

		if len(listZ) == 0:
			z = 1
		else:
			maxZinThisYear = max(listZ)
			z = maxZinThisYear + 1

		for i in range(posCount):
			posNumber 	= ws['A{}'.format(r)].value
			name 			= ws['B{}'.format(r)].value
			count 		= ws['C{}'.format(r)].value
			document 	= ws['D{}'.format(r)].value
			haract 		= ws['E{}'.format(r)].value
			acception 	= ws['F{}'.format(r)].value
			costKind 	= ws['G{}'.format(r)].value
			datestuff 	= datetime.today().strftime('%Y-%m-%d-%H:%M:%S')
			str = [id, '{}-{}'.format(z, year), date, theme, posNumber, name, count, document, haract, acception, costKind, makerD, makerFio, dopInfo, datestuff]
			wsr.append(str)
			for cell in wsr[wsr.max_row]:
				cell.style = "TabledDB"
			id += 1
			r  += 1
		print('База успешно дополнена')

	# если база ещё не создана, создаём её, вставляем шапку, дописываем заявку
	except FileNotFoundError:
		wbr = Workbook()
		wsr = wbr.active
		wsr.title = "База заявок"

		head = ['id', '№ заявки', 'Дата', 'Тема', '№ п/п', 'Наименование', 'Кол-во', "Нормативные документы", 'Технические хар-ки', 'Применение', 'Вид затрат', 'Должность составителя', 'Фамилия составителя', 'Дополнительные сведения', 'Дата внесения в базу'] 
		z 		= 1
		r 		= 13
		id 	= 1
		date 	= ws['A7'].value
		year 	= '20' + date[-3:-1]
		theme = ws['A9'].value
		makerD = 	ws['A{}'.format(18+posCount)].value
		makerFio = 	ws['F{}'.format(18+posCount)].value
		dopInfo 	= ws['A{}'.format(14+posCount)].value
		wsr.append(head)
		for cell in wsr[wsr.max_row]:
				cell.style = TabledDB

		for i in range(posCount):
			posNumber 	= ws['A{}'.format(r)].value
			name 			= ws['B{}'.format(r)].value
			count 		= ws['C{}'.format(r)].value
			document 	= ws['D{}'.format(r)].value
			haract 		= ws['E{}'.format(r)].value
			acception 	= ws['F{}'.format(r)].value
			costKind 	= ws['G{}'.format(r)].value
			datestuff = datetime.today().strftime('%Y-%m-%d-%H:%M:%S')
			str = [id, '{}-{}'.format(z, year), date, theme, posNumber, name, count, document, haract, acception, costKind, makerD, makerFio, dopInfo, datestuff]
			wsr.append(str)
			for cell in wsr[wsr.max_row]:
				cell.style = TabledDB
			id += 1
			r  += 1
		print('База успешно создана')

	# сохраняем и закрываем базу
	try:
		wbr.save(r'{}\bdz.xlsx'.format(cwd))
	except PermissionError:
		print('БД обновить не удалось, ибо к ней нет доступа. Быть может она открыта в экселе? Закройте её и повторите попытку!')
	wbr.close

def genZajavka():

	nz = input('Введите номер заявки из БД: ')

	# читаем базу
	wb_bd = load_workbook(filename='bdz.xlsx', data_only=True)
	ws_bd = wb_bd.active

	# выбираем данные из базы для генерации заявки
	arr = []
	for row in ws_bd:
		if row[1].value == nz:
			arr.append([cell.value for cell in row])
	if len(arr) == 0:
		return print('В базе отсутствует заявка с номером "{}".'.format(nz))

	# создаём новую заявку
	wbt = Workbook()
	wst = wbt.active

	# заполняем заявку
	wst.append(['','','','','','',''])
	wst.append(['','','','','','',''])
	wst.append(['','','','','','',''])
	try:
		img = Image('TTlogo.png')
		wst.add_image(img, 'A1')
	except FileNotFoundError:
		print('Картинка с логотипом Турботехники в данной папке отсутствует. Вставьте её в заявку вручную.')

	wst['A4'].value = 'Заявка на закупку в ОВК'
	wst['A5'].value = 'Лаборатория испытаний  ИЦ'
	wst['A6'].value = 'Наименование подразделения заказчика'
	wst['A7'].value = arr[0][2]
	wst['A8'].value = 'Дата подачи заявки'
	wst['A9'].value = arr[0][3]
	wst['A10'].value = 'Тема'
	wst.append(['','','','','','',''])
	wst['E4'].value = '''Утверждаю
	Генеральный директор
	АО «Турбокомплект»

	______________А.В.Барбалат
	«_____»_______________{}г.'''.format(nz[-4:])
		
	wst.append(['№ п/п','Наименование (обозначение)','Кол-во','Нормативные документы (ГОСТ, ТУ, ТЗ, чертежи и др.)','Технические хар-ки (параметры)','Применение','Вид затрат'])
	for cell in wst[wst.max_row]:
		cell.style = TabledC

	rt = 13
	for row in arr:
		wst['A{}'.format(rt)] = row[4]
		wst['B{}'.format(rt)] = row[5]
		wst['C{}'.format(rt)] = row[6]
		wst['D{}'.format(rt)] = row[7]
		wst['E{}'.format(rt)] = row[8]
		wst['F{}'.format(rt)] = row[9]
		wst['G{}'.format(rt)] = row[10]
		rt +=1

	wst.append(["Дополнительные сведения:"])
	wst.merge_cells('A{0}:G{0}'.format(13+len(arr)))
	wst['A{}'.format(13+len(arr))].border = Border(bottom=Side(border_style=None, color='FFFFFF'))
	wst['A{}'.format(14+len(arr))] = arr[0][13] #dop info
	wst.merge_cells('A{0}:G{0}'.format(14+len(arr)))

	wst.append(['','','','','','',''])
	wst.append(['','','','','','',''])

	wst['A{}'.format(17+len(arr))].value = 'Заявку составил:'
	wst['A{}'.format(18+len(arr))].value = arr[0][11] 			# должность составителя
	wst['F{}'.format(18+len(arr))].value = arr[0][12]			# фамилия составителя
	wst.merge_cells('F{0}:G{0}'.format(18+len(arr)))

	wst.append(['Согласовано:','','','','','',''])
	wst.row_dimensions[19+len(arr)].height = 33

	wst['A{}'.format(20+len(arr))].value = 'Руководитель подразделения'
	wst['F{}'.format(20+len(arr))].value = 'Зайцев О.Г.'
	wst.merge_cells('F{0}:G{0}'.format(20+len(arr)))
	wst.row_dimensions[20+len(arr)].height = 20

	wst['A{}'.format(21+len(arr))].value = 'Технический директор'
	wst['F{}'.format(21+len(arr))].value = 'Каминский Р.В.'
	wst.merge_cells('F{0}:G{0}'.format(21+len(arr)))
	wst.row_dimensions[21+len(arr)].height = 25

	wst.merge_cells('A4:D4')
	wst.merge_cells('A5:D5')
	wst.merge_cells('A6:D6')
	wst.merge_cells('A7:D7')
	wst.merge_cells('A8:D8')
	wst.merge_cells('A9:D9')
	wst.merge_cells('A10:D10')
	wst.merge_cells('E4:G9')

	k = 1
	wst.column_dimensions['A'].width = 5 * k
	wst.column_dimensions['B'].width = 16 * k
	wst.column_dimensions['C'].width = 8 * k
	wst.column_dimensions['D'].width = 17 * k
	wst.column_dimensions['E'].width = 16 * k
	wst.column_dimensions['F'].width = 13 * k
	wst.column_dimensions['G'].width = 7 * k

	wst.row_dimensions[9].height = 30

	# применим стили
	wst['A4'].style = fieldValue
	wst['A5'].style = fieldValue
	wst['A7'].style = fieldValue
	wst['A9'].style = fieldValue
	wst['E4'].style = alignR
	for cell in wst['A6:D6'][0]:
		cell.style = notice
	for cell in wst['A8:D8'][0]:
		cell.style = notice
	for cell in wst['A10:D10'][0]:
		cell.style = notice

	for row in wst['A13':'G{}'.format(12+len(arr))]:
		for cell in row:
			cell.style = TabledC

	for row in wst['A{}'.format(13+len(arr)):'G{}'.format(14+len(arr))]:
		for cell in row:
			cell.style = TabledL

	# стили для текста снизу заявки
	for row in wst['A{}'.format(17+len(arr)):'G{}'.format(21+len(arr))]:
		for cell in row:
			cell.style = alignL

	# сохраняем сгенерированную заявку
	wbt.save('заявка_№{}.xlsx'.format(nz))
	wbt.close
	print('Заявка успешно сгенерирована')

def styles():
	bdr_thin = Side(style='thin', color="000000")

	# стиль "TabledDB"
	global TabledDB
	TabledDB = NamedStyle(name="TabledDB")
	TabledDB.border = Border(left=bdr_thin, top=bdr_thin, right=bdr_thin, bottom=bdr_thin)
	TabledDB.alignment = Alignment(vertical = 'center', horizontal = 'center')
	TabledDB.font = Font(name='Times New Roman', size=12)

	# стиль "TabledC"
	global TabledC
	TabledC = NamedStyle(name="TabledC")
	TabledC.border = Border(left=bdr_thin, top=bdr_thin, right=bdr_thin, bottom=bdr_thin)
	TabledC.alignment = Alignment(wrap_text = 'True', vertical = 'center', horizontal = 'center')
	TabledC.font = Font(name='Times New Roman', size=12)

	# стиль "TabledL"
	global TabledL
	TabledL = NamedStyle(name="TabledL")
	TabledL.border = Border(left=bdr_thin, top=bdr_thin, right=bdr_thin, bottom=bdr_thin)
	TabledL.alignment = Alignment(wrap_text = 'True', vertical = 'center', horizontal = 'left')
	TabledL.font = Font(name='Times New Roman', size=12)

	# создадим стиль "underscore"
	global underscore
	underscore = NamedStyle(name="underscore")
	underscore.border = Border(bottom=bdr_thin)
	underscore.alignment = Alignment(vertical = 'bottom', horizontal = 'left')
	underscore.font = Font(name='Times New Roman', size=12)

	# создадим стиль "alignL"
	global alignL
	alignL = NamedStyle(name="alignL")
	alignL.alignment = Alignment(vertical = 'bottom', horizontal = 'left')
	alignL.font = Font(name='Times New Roman', size=12)

	# создадим стиль "alignR"
	global alignR
	alignR = NamedStyle(name="alignR")
	alignR.alignment = Alignment(wrap_text = 'True', vertical = 'top', horizontal = 'right')
	alignR.font = Font(name='Times New Roman', size=12)

	# стиль "fieldValue"
	global fieldValue
	fieldValue = NamedStyle(name="fieldValue")
	fieldValue.alignment = Alignment(vertical = 'bottom', horizontal = 'center')
	fieldValue.font = Font(name='Times New Roman', size=12)

	# стиль "notice"
	global notice
	notice = NamedStyle(name="notice")
	notice.border = Border(top=bdr_thin)
	notice.alignment = Alignment(vertical = 'top', horizontal = 'center')
	notice.font = Font(name='Times New Roman', size=8)

# body
warnings.filterwarnings("ignore")
cwd = getcwd()
print('='*45)
print('Эта программа для управления заявками и их БД.')
styles()

while True:
	action = input('''Выберите действие:
1 Дополнить базу данных заявок;
2 Сгенерировать заявку по её номеру в БД;
3 Выход.
->''')
	if action == '1':
		db_filler()
	elif action == '2':
		genZajavka()
	elif action == '3':
		print('Работа завершена')
		break
	else:
		print("Вы заблудились в трёх соснах")
	print('-'*45)